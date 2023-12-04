from openpyxl import Workbook, load_workbook

from win32com import client

app = client.DispatchEx("Excel.Application")

app.Interactive = False

app.Visible = False

path = input("Coloque o caminho do arquivo: ")

workbook = app.Workbooks.Open(path)
workbook.ActiveSheet.ExportAsFixedFormat(0, path)
workbook.Close()

from datetime import datetime, timedelta

def gerador_folhas_col(dados, modelo_folha_ponto):
    planilha_nomes = load_workbook(dados)

    aba_ativa_nomes = planilha_nomes.active

    for celula in aba_ativa_nomes["A"]:
        planilha_folha_ponto = load_workbook(modelo_folha_ponto)
        aba_ativa_folha = planilha_folha_ponto.active

        linha_nome = celula.row
        setor = aba_ativa_nomes[f"D{linha_nome}"].value
        nome = aba_ativa_nomes[f"B{linha_nome}"].value
        cargo = aba_ativa_nomes[f"C{linha_nome}"].value
        mat = aba_ativa_nomes[f"A{linha_nome}"].value

        aba_ativa_folha["C5"] = nome
        aba_ativa_folha["C6"] = cargo
        aba_ativa_folha["G6"] = mat

        nome_planilha = str(f"{setor}_{nome}.xlsx")
        planilha_folha_ponto.save(nome_planilha)
        path = f"C:\Users\usuario\OneDrive\Documentos\GitHub\Automatizacao-folha-de-ponto\arquivos\testes com arquivos modelo\{setor}_{nome}.xlsx"
        planilha_folha_ponto.ExportAsFixedFormat(0, path)



"""def gerador_folha_mensal(modelo_folha_ponto):
    print("Teste")

meses = {"jan":("01/01/2024")}"""
gerador_folhas_col("modelo-dados-teste.xlsx", "modelo-tabela-folha-ponto-teste.xlsx")

#gerador_modelo_folha("modelo-tabela-ponto-teste.xlsx")

