import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

# Ler a planilha original com Pandas
data = pd.read_excel('modelo-dados-teste.xlsx')

# Carregar o modelo de folha de ponto em Excel
workbook = load_workbook('modelo-tabela-folha-ponto-teste.xlsx')
sheet = workbook.active

# Preencher o modelo com os dados da planilha original
for index, row in data.iterrows():
    matricula = row['MATRICULA']
    nome = row['COLABORADOR']
    funcao = row['CARGO']

    # Preencher as células do modelo com os dados
    sheet.cell(row=index + 2, column=1).value = matricula
    sheet.cell(row=index + 2, column=2).value = nome
    sheet.cell(row=index + 2, column=3).value = funcao

# Salvar o arquivo Excel preenchido
workbook.save('folha_de_ponto_preenchida.xlsx')

# Aqui você pode adicionar a lógica para converter o arquivo Excel para PDF, utilizando a biblioteca de sua escolha.
# Por exemplo, usando 'xlsx2pdf' ou 'PyPDF2'.
