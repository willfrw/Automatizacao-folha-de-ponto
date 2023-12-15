import os
from openpyxl import Workbook, load_workbook
import win32com.client
from datetime import datetime, timedelta


dia_semana_dict = {
    'Sunday': 'DOMINGO',
    'Monday': 'SEGUNDA-FEIRA',
    'Tuesday': 'TERÇA-FEIRA',
    'Wednesday': 'QUARTA-FEIRA',
    'Thursday': 'QUINTA-FEIRA',
    'Friday': 'SEXTA-FEIRA',
    'Saturday': 'SÁBADO'
}

mes_dict = {
    '1': 'JANEIRO',
    '2': 'FEVEREIRO',
    '3': 'MARÇO',
    '4': 'ABRIL',
    '5': 'MAIO',
    '6': 'JUNHO',
    '7': 'JULHO',
    '8': 'AGOSTO',
    '9': 'SETEMBRO',
    '10': 'OUTUBRO',
    '11': 'NOVEMBRO',
    '12': 'DEZEMBRO'
}

feriado_list = [
    '01/01/2024',  # Ano Novo
    '13/02/2024',  # Carnaval
    '29/03/2024',  # Sexta-Feira Santa
    '21/04/2024',  # Dia de Tiradentes
    '23/04/2024',  # Feriado Municipal
    '01/05/2024',  # Dia do Trabalho
    '26/05/2024',  # Feriado Municipal
    '30/05/2024',  # Corpus Christi
    '15/08/2024',  # Feriado Municipal
    '07/09/2024',  # Independência do Brasil
    '12/10/2024',  # Nossa Senhora Aparecida
    '15/10/2024',  # Dia do Professor
    '28/10/2024',  # Dia do Servidor Público
    '02/11/2024',  # Dia de Finados
    '15/11/2024',  # Proclamação da República
    '20/11/2024',  # Dia da Consciência Negra
    '25/12/2024'  # Natal'
]

coluna_list = ['E', 'F', 'G', 'H', 'I']

qnt_dias_mes = []

def gerador_folhas_col(dados, modelo_folha_ponto):
    planilha_nomes = load_workbook(dados)
    aba_ativa_nomes = planilha_nomes.active

    for celula in aba_ativa_nomes["A"]:

        linha_nome = celula.row

        if linha_nome > 1:
            setor_cod = aba_ativa_nomes[f"D{linha_nome}"].value
            setor_nome = aba_ativa_nomes[f"E{linha_nome}"].value
            nome = aba_ativa_nomes[f"B{linha_nome}"].value
            cargo = aba_ativa_nomes[f"C{linha_nome}"].value
            mat = aba_ativa_nomes[f"A{linha_nome}"].value
            if cargo != 'CARGO2':
                planilha_folha_ponto = load_workbook(modelo_folha_ponto[0])
                aba_ativa_folha = planilha_folha_ponto.active
                aba_ativa_folha["C4"] = "NOME: " + str(nome)
                aba_ativa_folha["C5"] = "CARGO: " + str(cargo)
                aba_ativa_folha["G5"] = "MATRÍCULA: " + str(mat)

                nome_planilha = str(f"{setor_cod}_{nome}.xlsx")
            else:
                planilha_folha_ponto = load_workbook(modelo_folha_ponto[1])
                aba_ativa_folha = planilha_folha_ponto.active
                aba_ativa_folha["C4"] = "NOME: " + str(nome)
                aba_ativa_folha["C5"] = "CARGO: " + str(cargo)
                aba_ativa_folha["G5"] = "MATRÍCULA: " + str(mat)

                nome_planilha = str(f"{setor_cod}_{nome}.xlsx")

            planilha_folha_ponto.save(nome_planilha)
    return 0

# A função irá criar um arquivo em pdf com base no excel
# Criar um loop for com o tamanho da lista com os arquivo para criar conforme percorsse o diretório
def gerador_pdf(path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    workbook = excel.Workbooks.Open(path)
    path_pdf = path.replace(".xlsx", ".pdf")

    try:
        workbook.ActiveSheet.ExportAsFixedFormat(0, path_pdf)
    except Exception as e:
        print(f"Erro ao exportar {path} para PDF: {str(e)}")
    finally:
        workbook.Close(False)
        excel.Quit()

    return 0


# Cria o modelo de folha com base no mês, sendo esta utilizada para vincular aos nomes dos funcionários
def gerador_folha_mensal(mes, ano, modelo, lista_nome_gerado=[2]):
    for m in range(1, 13):
        if ((m) % 2 != 0 and (m) <= 7) or ((m) % 2 == 0 and (m) >= 8):
            qnt_dias_mes.append(31)
        elif (m) == 2:
            if (ano % 4 == 0 and ano % 100 != 0) or (ano % 400 == 0):
                qnt_dias_mes.append(29)
            else:
                qnt_dias_mes.append(28)
        else:
            qnt_dias_mes.append(30)

    n_mes = datetime(year=ano, month=mes, day=1)

    '''for i in range(0, qnt_dias_mes[mes-1]):
        if dia_semana_dict[n_mes.strftime("%A")] == 'DOMINGO' or dia_semana_dict[n_mes.strftime("%A")] == 'SÁBADO':
            print(f'{n_mes.strftime("%d/%m/%Y")} {dia_semana_dict[n_mes.strftime("%A")]} {dia_semana_dict[n_mes.strftime("%A")]} {dia_semana_dict[n_mes.strftime("%A")]}')
        else:
            print(f'{n_mes.strftime("%d/%m/%Y")} {dia_semana_dict[n_mes.strftime("%A")]}')
        n_mes += timedelta(days=1)'''

    folha_de_ponto = load_workbook(modelo)
    folha_de_ponto_prt = load_workbook(modelo)

    aba_folha_de_ponto = folha_de_ponto.active
    aba_folha_de_ponto_prt = folha_de_ponto_prt.active

    aba_folha_de_ponto['C3'] = f"PERÍODO: 01 A {qnt_dias_mes[mes - 1]} DE {mes_dict[str(mes)]} DE {ano}"
    aba_folha_de_ponto_prt['C3'] = f"PERÍODO: 01 A {qnt_dias_mes[mes - 1]} DE {mes_dict[str(mes)]} DE {ano}"

    for celula in aba_folha_de_ponto['C']:
        linha = celula.row
        if linha >= 7 and linha <= qnt_dias_mes[mes - 1] + 6:
            #print(qnt_dias_mes[mes - 1])
            aba_folha_de_ponto[f'C{linha}'] = f'{n_mes.strftime("%d/%m/%Y")}'
            aba_folha_de_ponto[f'D{linha}'] = f'{dia_semana_dict[n_mes.strftime("%A")]}'
            if dia_semana_dict[n_mes.strftime("%A")] == 'SÁBADO' or dia_semana_dict[n_mes.strftime("%A")] == 'DOMINGO':
                for letra in range(0, len(coluna_list)):
                    aba_folha_de_ponto[f'{coluna_list[letra]}{linha}'] = f'{dia_semana_dict[n_mes.strftime("%A")]}'
            elif n_mes.strftime("%d/%m/%Y") in feriado_list:
                for letra in range(0, len(coluna_list)):
                    aba_folha_de_ponto[f'{coluna_list[letra]}{linha}'] = f'FERIADO'
            #print(aba_folha_de_ponto[f'C{linha}'].value)
            n_mes += timedelta(days=1)

    n_mes = datetime(year=ano, month=mes, day=1)

    for celula in aba_folha_de_ponto_prt['C']:
        linha = celula.row
        if linha >= 7 and linha <= qnt_dias_mes[mes - 1] + 6:
            #print(qnt_dias_mes[mes - 1])
            aba_folha_de_ponto_prt[f'C{linha}'] = f'{n_mes.strftime("%d/%m/%Y")}'
            aba_folha_de_ponto_prt[f'D{linha}'] = f'{dia_semana_dict[n_mes.strftime("%A")]}'
            #print(aba_folha_de_ponto_prt[f'C{linha}'].value)
            n_mes += timedelta(days=1)

    nome_folha = f'{str(mes)} - {mes_dict[str(mes)]} {str(ano)}_dne.xlsx'
    nome_folha_prt = f'{str(mes)} - {mes_dict[str(mes)]} {str(ano)} - PORT_dne.xlsx'
    lista_nome_gerado.append(nome_folha)
    lista_nome_gerado.append(nome_folha_prt)
    folha_de_ponto.save(nome_folha)
    folha_de_ponto_prt.save(nome_folha_prt)


"""def gerador_folha_mensal(modelo_folha_ponto):
    print("Teste")

meses = {"jan":("01/01/2024")}"""




