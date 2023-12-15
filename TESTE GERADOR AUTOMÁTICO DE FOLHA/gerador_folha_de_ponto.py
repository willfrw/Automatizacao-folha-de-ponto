from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from tkinter.filedialog import askopenfilename, askdirectory

dia_semana_dict = {
    'Sunday': 'DOMINGO',
    'Monday': 'SEGUNDA-FEIRA',
    'Tuesday': 'TERÇA-FEIRA',
    'Wednesday': 'QUARTA-FEIRA',
    'Thursday': 'QUINTA-FEIRA',
    'Friday': 'SEXTA-FEIRA',
    'Saturday': 'SÁBADO'
}

mes_dict = {
    '1': 'JANEIRO',
    '2': 'FEVEREIRO',
    '3': 'MARÇO',
    '4': 'ABRIL',
    '5': 'MAIO',
    '6': 'JUNHO',
    '7': 'JULHO',
    '8': 'AGOSTO',
    '9': 'SETEMBRO',
    '10': 'OUTUBRO',
    '11': 'NOVEMBRO',
    '12': 'DEZEMBRO'
}

feriado_list = [
    '01/01/2024',  # Ano Novo
    '13/02/2024',  # Carnaval
    '29/03/2024',  # Sexta-Feira Santa
    '21/04/2024',  # Dia de Tiradentes
    '23/04/2024',  # Feriado Municipal
    '01/05/2024',  # Dia do Trabalho
    '26/05/2024',  # Feriado Municipal
    '30/05/2024',  # Corpus Christi
    '15/08/2024',  # Feriado Municipal
    '07/09/2024',  # Independência do Brasil
    '12/10/2024',  # Nossa Senhora Aparecida
    '15/10/2024',  # Dia do Professor
    '28/10/2024',  # Dia do Servidor Público
    '02/11/2024',  # Dia de Finados
    '15/11/2024',  # Proclamação da República
    '20/11/2024',  # Dia da Consciência Negra
    '25/12/2024'  # Natal'
]

coluna_list = ['E', 'F', 'G', 'H', 'I']

qnt_dias_mes = []

modelo = askopenfilename(title="Planilha Modelo")
pasta_origem = askdirectory(title="Pasta Origem")
pasta_destino = askdirectory(title="Pasta Destino")
mes = int(input("Digite o mês: "))
ano = int(input("Digite o ano: "))


def gerador_folha_mensal(mes, ano, modelo):
    for m in range(1, 13):
        if ((m) % 2 != 0 and (m) <= 7) or ((m) % 2 == 0 and (m) >= 8):
            qnt_dias_mes.append(31)
        elif (m) == 2:
            if (ano % 4 == 0 and ano % 100 != 0) or (ano % 400 == 0):
                qnt_dias_mes.append(29)
            else:
                qnt_dias_mes.append(28)
        else:
            qnt_dias_mes.append(30)

    n_mes = datetime(year=ano, month=mes, day=1)

    '''for i in range(0, qnt_dias_mes[mes-1]):
        if dia_semana_dict[n_mes.strftime("%A")] == 'DOMINGO' or dia_semana_dict[n_mes.strftime("%A")] == 'SÁBADO':
            print(f'{n_mes.strftime("%d/%m/%Y")} {dia_semana_dict[n_mes.strftime("%A")]} {dia_semana_dict[n_mes.strftime("%A")]} {dia_semana_dict[n_mes.strftime("%A")]}')
        else:
            print(f'{n_mes.strftime("%d/%m/%Y")} {dia_semana_dict[n_mes.strftime("%A")]}')
        n_mes += timedelta(days=1)'''

    folha_de_ponto = load_workbook(modelo)
    folha_de_ponto_prt = load_workbook(modelo)

    aba_folha_de_ponto = folha_de_ponto.active
    aba_folha_de_ponto_prt = folha_de_ponto_prt.active

    aba_folha_de_ponto['C3'] = f"PERÍODO: 01 A {qnt_dias_mes[mes - 1]} DE {mes_dict[str(mes)]} DE {ano}"
    aba_folha_de_ponto_prt['C3'] = f"PERÍODO: 01 A {qnt_dias_mes[mes - 1]} DE {mes_dict[str(mes)]} DE {ano}"

    for celula in aba_folha_de_ponto['C']:
        linha = celula.row
        if linha >= 7 and linha <= qnt_dias_mes[mes - 1] + 6:
            #print(qnt_dias_mes[mes - 1])
            aba_folha_de_ponto[f'C{linha}'] = f'{n_mes.strftime("%d/%m/%Y")}'
            aba_folha_de_ponto[f'D{linha}'] = f'{dia_semana_dict[n_mes.strftime("%A")]}'
            if dia_semana_dict[n_mes.strftime("%A")] == 'SÁBADO' or dia_semana_dict[n_mes.strftime("%A")] == 'DOMINGO':
                for letra in range(0, len(coluna_list)):
                    aba_folha_de_ponto[f'{coluna_list[letra]}{linha}'] = f'{dia_semana_dict[n_mes.strftime("%A")]}'
            elif n_mes.strftime("%d/%m/%Y") in feriado_list:
                for letra in range(0, len(coluna_list)):
                    aba_folha_de_ponto[f'{coluna_list[letra]}{linha}'] = f'FERIADO'
            #print(aba_folha_de_ponto[f'C{linha}'].value)
            n_mes += timedelta(days=1)

    n_mes = datetime(year=ano, month=mes, day=1)

    for celula in aba_folha_de_ponto_prt['C']:
        linha = celula.row
        if linha >= 7 and linha <= qnt_dias_mes[mes - 1] + 6:
            print(qnt_dias_mes[mes - 1])
            aba_folha_de_ponto_prt[f'C{linha}'] = f'{n_mes.strftime("%d/%m/%Y")}'
            aba_folha_de_ponto_prt[f'D{linha}'] = f'{dia_semana_dict[n_mes.strftime("%A")]}'
            print(aba_folha_de_ponto_prt[f'C{linha}'].value)
            n_mes += timedelta(days=1)

    nome_folha = f'{str(mes)} - {mes_dict[str(mes)]} {str(ano)}.xlsx'
    nome_folha_prt = f'{str(mes)} - {mes_dict[str(mes)]} {str(ano)} - PORT.xlsx'

    folha_de_ponto.save(nome_folha)
    folha_de_ponto_prt.save(nome_folha_prt)

gerador_folha_mensal(mes, ano, modelo)